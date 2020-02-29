#!/usr/bin/env python

import requests
from lxml import html
import re
import openpyxl

filename_read = 'MotoBudrex_STOCK_KROPEK.XLSX'
filename_write = 'MotoBudrex_STOCK_KROPEK_allegro.XLSX'
col_start = 4
col_stop = 4
row_start = 2
row_stop = 53
opony = []
ceny = []
ceny_temp = []
ilosci = []
ilosci_temp = []
popularnosc = []
popularnosc_temp = []

#reading form xls
def read_xls(filename_read):
    global opony
    print("Otwieram plik", filename_read)
    wb = openpyxl.load_workbook(filename=filename_read)
    ws = wb.active
    for col in ws.iter_cols(min_col=col_start, max_col=col_stop, min_row=row_start, max_row=row_stop):
        for cell in col:
            print(cell.value)
            opony.append(cell.value)
    print(opony)
    wb.close()
    return opony

def allegropl_price_check(opony):
    global ceny
    ceny_temp = []
    global popularnosc
    popularnosc_temp = []
    for i in range(len(opony)):
        print("Sprawdzanie cen "+ str(i+1) +" z "+ str(len(opony)) +": "+ str(opony[i]))
        # pobranie zawartosci strony
        
        # wyszukiwanie przez wyszukiwarkę allegro
        # page = requests.get("https://allegro.pl/listing?string="+nagrody[i]+"&order=d")
        
        # opony, komplet
        page = requests.get("https://allegro.pl/kategoria/opony-do-samochodow-osobowych-257688?liczba-opon-w-ofercie=Komplet%204%20szt.&string="+opony[i]+"&order=qd&bmatch=baseline-cl-dict43-aut-1-4-1127&stan=nowe")

        # opony, dwie sztuki
        # page = requests.get("https://allegro.pl/kategoria/opony-do-samochodow-osobowych-257688?liczba-opon-w-ofercie=2%20szt.&string="+opony[i]+"&order=qd&bmatch=baseline-cl-dict43-aut-1-4-1127&stan=nowe")

        tree = html.fromstring(page.content)

        # tworzenie listy ile osob kupilo
        popularnosc_temp.append(tree.xpath('//span[@class ="_9c44d_2o04k"]/text()'))
        popularnosc_temp = sum(popularnosc_temp, [])
        popularnosc_temp = [popular.replace(' osoba kupiła', '') for popular in popularnosc_temp]
        popularnosc_temp = [popular.replace(' osoby kupiły', '') for popular in popularnosc_temp]
        popularnosc_temp = [popular.replace(' osób kupiło', '') for popular in popularnosc_temp]
        popularnosc_temp = [popular.replace(' osoba licytuje', '') for popular in popularnosc_temp]
        popularnosc_temp = [popular.replace(' osoby licytują', '') for popular in popularnosc_temp]
        popularnosc_temp = [popular.replace(' osób licytuje', '') for popular in popularnosc_temp]
        popularnosc_temp = [popular.replace('nikt nie licytuje', '0') for popular in popularnosc_temp]
        popularnosc_temp = [int(popular) for popular in popularnosc_temp]
        popularnosc_temp = sum(popularnosc_temp)
        print("Ilosc kupionych: "+str(popularnosc_temp))
        popularnosc.append(popularnosc_temp)
        popularnosc_temp = []

        # tworzenie listy cen
        ceny_temp.append(tree.xpath('//span[@class="_9c44d_1zemI"]/text()'))
        ceny_temp = sum(ceny_temp, [])
        if not ceny_temp:
            ceny_temp = '0'
        ceny_temp = [cena.replace(' ', '') for cena in ceny_temp]
        ceny_temp = [int(cena) for cena in ceny_temp]
        print("Cena minimalna: "+str(min(ceny_temp)))
        ceny.append(min(ceny_temp))
        ceny_temp = []
    return (ilosci, popularnosc, ceny)

#wrtie lists to xls
def write_xls(filename_write):
    print("Zapisuję do xls...")
    wb = openpyxl.load_workbook(filename=filename_read)
    ws = wb.active
    ws['U1'] = "Ilosc kupionych"
    ws['V1'] = "Cena min"
    i = row_start
    for n in popularnosc:
        ws['U' + str(i)] = n
        i = i + 1
    i = row_start
    for n in ceny:
        ws['V' + str(i)] = n
        i = i + 1
    wb.save(filename_write)
    print("OK")

read_xls(filename_read)
allegropl_price_check(opony)
write_xls(filename_write)
