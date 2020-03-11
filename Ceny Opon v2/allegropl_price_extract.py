#!/usr/bin/env python

import requests
from lxml import html
import openpyxl
import time
from random import randint

filename_read = 'Lista_aukcji_zamowienia_informacje.XLSX'
filename_write = 'Lista_aukcji_zamowienia_informacje_updated.XLSX'
row_start = 2
row_stop = 120

col_start = 1
col_stop = 1
opony = []
prom_popularnosc = []
prom_cena_min = []
prom_cena_dost = []
prom_cena_max = {}
regular_popularnosc = []
regular_cena_min = []
regular_cena_dost = []
regular_cena_max = {}


# pobieranie wszystkich opon z pliku,
# wyszukiwanie wszystkich opon na allegro jedna po drugiej i zapisywanie do pamieci,
# zapisywanie wszystkiego do pliku,
# w przypadku bledu zapisanie biezacej zawartosci pamieci do pliku

#reading form xls
def read_xls(filename_read):
   # global opony
    i = 1
    print("Otwieram plik", filename_read)
    wb = openpyxl.load_workbook(filename=filename_read)
    ws = wb.active
    for col in ws.iter_cols(min_col=col_start, max_col=col_stop, min_row=row_start, max_row=row_stop):
        for cell in col:
            print(i, cell.value)
            opony.append(cell.value)
            i = i+1
    wb.close()
    return opony


def allegropl_price_check(opony):
    cena_min_temp = []
    cena_dost_temp = []
    cena_max_temp = {}
    popularnosc_temp = []
    for i in range(len(opony)):
        print("Sprawdzanie cen "+ str(i+1) + " z " + str(len(opony)) + ": " + str(opony[i]))
        # pobranie zawartosci strony
        
        # wyszukiwanie przez wyszukiwarkę allegro
        # page = requests.get("https://allegro.pl/listing?string="+nagrody[i]+"&order=d")
        # opony, komplet
        page = requests.get("https://allegro.pl/kategoria/opony-do-samochodow-osobowych-257688?liczba-opon-w-ofercie=Komplet%204%20szt.&string="+opony[i]+"&order=qd&bmatch=baseline-cl-dict43-aut-1-4-1127&stan=nowe")
        # opony, dwie sztuki
        # page = requests.get("https://allegro.pl/kategoria/opony-do-samochodow-osobowych-257688?liczba-opon-w-ofercie=2%20szt.&string="+opony[i]+"&order=qd&bmatch=baseline-cl-dict43-aut-1-4-1127&stan=nowe")

        tree = html.fromstring(page.content)

        # tworzenie listy popularnosci - promowane
        try:
            popularnosc_temp.append(tree.xpath('//div/section[2]/section//span[@class="_9c44d_2o04k"]/text()'))
            popularnosc_temp = sum(popularnosc_temp, [])
            popularnosc_temp = [popular.replace(' osoba kupiła', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace(' osoby kupiły', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace(' osób kupiło', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace(' osoba licytuje', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace(' osoby licytują', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace(' osób licytuje', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace('nikt nie licytuje', '0') for popular in popularnosc_temp]
            popularnosc_temp = [int(popular) for popular in popularnosc_temp]
            print("-- OFERTY PROMOWANE")
            #popularnosc_temp = sum(popularnosc_temp)
            print("Ilosc kupionych: "+str(sum(popularnosc_temp)))
            prom_popularnosc.append(sum(popularnosc_temp))
            popularnosc_temp.clear()
        except Exception as error:
            print(time.ctime(time.time()), error)
            prom_popularnosc.append(str(error))
            pass

        # tworzenie listy cen minimalnych - promowane
        try:
            cena_min_temp.append(tree.xpath('//div/section[2]/section//span[@class="_9c44d_1zemI"]/text()'))
            cena_min_temp= sum(cena_min_temp, [])
            if not cena_min_temp:
                cena_min_temp = '0'
            cena_min_temp = [cena.replace(' ', '') for cena in cena_min_temp]
            cena_min_temp = [int(cena) for cena in cena_min_temp]
            print("Cena minimalna: "+str(min(cena_min_temp)))
            prom_cena_min.append(min(cena_min_temp))
            cena_min_temp.clear()
        except Exception as error:
            print(time.ctime(time.time()), error)
            prom_cena_min.append(str(error))
            pass

        # tworzenie listy cen z dostawą - promowane
        try:
            cena_dost_temp.append(tree.xpath('//div/section[2]//span[@class="_9c44d_1zemI"]/text()|//div/section[2]//div[@class="_9c44d_1xKGX"]//i/text()'))
            cena_dost_temp = sum(cena_dost_temp, [])
            for j in range(len(cena_dost_temp)):
                if 'zł' in cena_dost_temp[j]:
                    k = j - 1
                    cena_dost_temp.pop(k)
                    cena_dost_temp.append('usuniety')
            cena_dost_temp = [cena for cena in cena_dost_temp if not ' z dostawą' in cena]
            cena_dost_temp = [cena for cena in cena_dost_temp if not 'darmowa' in cena]
            cena_dost_temp = [cena for cena in cena_dost_temp if not 'dostawa' in cena]
            cena_dost_temp = [cena for cena in cena_dost_temp if not 'usuniety' in cena]
            cena_dost_temp = [cena[:-6] if 'zł' in cena else cena for cena in cena_dost_temp]
            cena_dost_temp = [cena.replace(' ', '') for cena in cena_dost_temp]
            cena_dost_temp = [int(cena) for cena in cena_dost_temp]
            print("Cena z dostawą: "+str(min(cena_dost_temp)))
            prom_cena_dost.append(min(cena_dost_temp))
            cena_dost_temp.clear()
        except Exception as error:
            print(time.ctime(time.time()), error)
            prom_cena_dost.append(str(error))
            pass

        # # tworzenie listy cen max dla popularnosci >2 - promowane
        # cena_max_temp = dict(zip(cena_dost_temp, popularnosc_temp))
        # print(cena_max_temp)

        # tworzenie listy popularnosci - regularne
        try:
            popularnosc_temp.append(tree.xpath('//div/section[3]/section//span[@class="_9c44d_2o04k"]/text()'))
            popularnosc_temp = sum(popularnosc_temp, [])
            popularnosc_temp = [popular.replace(' osoba kupiła', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace(' osoby kupiły', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace(' osób kupiło', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace(' osoba licytuje', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace(' osoby licytują', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace(' osób licytuje', '') for popular in popularnosc_temp]
            popularnosc_temp = [popular.replace('nikt nie licytuje', '0') for popular in popularnosc_temp]
            popularnosc_temp = [int(popular) for popular in popularnosc_temp]
            print("-- OFERTY REGULARNE")
            #popularnosc_temp = sum(popularnosc_temp)
            print("Ilosc kupionych: "+str(sum(popularnosc_temp)))
            regular_popularnosc.append(sum(popularnosc_temp))
            popularnosc_temp.clear()
        except Exception as error:
            print(time.ctime(time.time()), error)
            regular_popularnosc(str(error))
            pass

        # tworzenie listy cen minimalnych - regularne
        try:
            cena_min_temp.append(tree.xpath('//div/section[3]/section//span[@class="_9c44d_1zemI"]/text()'))
            cena_min_temp= sum(cena_min_temp, [])
            if not cena_min_temp:
                cena_min_temp = '0'
            cena_min_temp = [cena.replace(' ', '') for cena in cena_min_temp]
            cena_min_temp = [int(cena) for cena in cena_min_temp]
            print("Cena minimalna: "+str(min(cena_min_temp)))
            regular_cena_min.append(min(cena_min_temp))
            cena_min_temp.clear()
        except Exception as error:
            print(time.ctime(time.time()), error)
            regular_cena_min.append(str(error))
            pass

        # tworzenie listy cen z dostawą - regularne
        try:
            cena_dost_temp.append(tree.xpath('//div/section[3]/section//span[@class="_9c44d_1zemI"]/text()|//div/section[3]//div[@class="_9c44d_1xKGX"]//i/text()'))
            cena_dost_temp = sum(cena_dost_temp, [])
            for j in range(len(cena_dost_temp)):
                if 'zł' in cena_dost_temp[j]:
                    k = j - 1
                    cena_dost_temp.pop(k)
                    cena_dost_temp.append('usuniety')
            cena_dost_temp = [cena for cena in cena_dost_temp if not ' z dostawą' in cena]
            cena_dost_temp = [cena for cena in cena_dost_temp if not 'darmowa' in cena]
            cena_dost_temp = [cena for cena in cena_dost_temp if not 'dostawa' in cena]
            cena_dost_temp = [cena for cena in cena_dost_temp if not 'usuniety' in cena]
            cena_dost_temp = [cena[:-6] if 'zł' in cena else cena for cena in cena_dost_temp]
            cena_dost_temp = [cena.replace(' ', '') for cena in cena_dost_temp]
            cena_dost_temp = [int(cena) for cena in cena_dost_temp]
            print("Cena z dostawą: "+str(min(cena_dost_temp)))
            regular_cena_dost.append(min(cena_dost_temp))
            cena_dost_temp.clear()
        except Exception as error:
            print(time.ctime(time.time()), error)
            regular_cena_dost.append(str(error))
            pass
    return (prom_popularnosc, prom_cena_min, prom_cena_dost, regular_popularnosc, regular_cena_min, regular_cena_dost)

#wrtie lists to xls
def write_xls(filename_write):
    print("Zapisuję do xls...")
    wb = openpyxl.load_workbook(filename=filename_read)
    ws = wb.active
    ws['M1'] = "Promowane - popularnosc"
    ws['N1'] = "Promowane - cena min"
    ws['O1'] = "Promowane - cena z dostawa"
    ws['X1'] = "Regularne - popularnosc"
    ws['Y1'] = "Regularne - cena min"
    ws['Z1'] = "Regularne - cena z dostawa"
    i = row_start
    for n in prom_popularnosc:
        ws['M' + str(i)] = n
        i = i + 1
    i = row_start
    for n in prom_cena_min:
        ws['N' + str(i)] = n
        i = i + 1
    i = row_start
    for n in prom_cena_dost:
        ws['O' + str(i)] = n
        i = i + 1
    i = row_start
    for n in regular_popularnosc:
        ws['X' + str(i)] = n
        i = i + 1
    i = row_start
    for n in regular_cena_min:
        ws['Y' + str(i)] = n
        i = i + 1
    i = row_start
    for n in regular_cena_dost:
        ws['Z' + str(i)] = n
        i = i + 1
    wb.save(filename_write)
    wb.close()
    print("OK")

try:
    read_xls(filename_read)
    try:
        allegropl_price_check(opony)
    except Exception as error:
        print(time.ctime(time.time()), error)
        pass
    write_xls(filename_write)
except Exception as error:
    print(time.ctime(time.time()), error)

